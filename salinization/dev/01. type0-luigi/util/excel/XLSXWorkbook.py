import logging
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException


class XLSXWorkbook:
    def __init__(self, filename: str):
        self.filename = filename

    @property
    def filename(self):
        return self.__filename

    @filename.setter
    def filename(self, filename: str, discard: bool = False):
        if not discard:  # save modified content back to the excel if needed
            try:
                if self.__workbook is not None and self.__dirty:
                    self.__workbook.save(self.filename)
            except AttributeError:
                pass

        # open the new excel
        try:
            self.__workbook = load_workbook(filename)
        except InvalidFileException as e:
            logging.error(f'Failed to open excel file {filename}: {e}')
            self.__workbook = None
        else:
            self.__filename = filename
        finally:
            self.__dirty = False

    @property
    def sheet_names(self) -> list:
        if self.__workbook is not None:
            return self.__workbook.sheetnames

        return None


if __name__ == '__main__':
    workbook = XLSXWorkbook('../../../../dataset/bentre/So Lieu Man Ben Tre 2018.xlsx')
    print(workbook.sheet_names)
